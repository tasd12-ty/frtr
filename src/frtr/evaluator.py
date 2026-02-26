"""Benchmark evaluator for FRTR-Bench.

Loads Questions sheets from all workbooks, runs the FRTR pipeline
on each question, and computes accuracy, latency, and token metrics
as defined in Section 4.2 of the paper.
"""

from __future__ import annotations

import io
import json
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from rich.console import Console
from rich.table import Table

from .config import FRTRConfig
from .reasoner import LLMReasoner, ReasoningResult
from .retriever import HybridRetriever

console = Console()


@dataclass
class QuestionRecord:
    """A single question from the benchmark."""

    workbook_name: str
    workbook_path: Path
    question: str
    reasoning_type: str
    expected_answer: str
    provenance: str
    difficulty: str


@dataclass
class EvalResult:
    """Result of evaluating a single question."""

    question: QuestionRecord
    predicted_answer: str
    reasoning: str
    correct: Optional[bool]  # None if manual review needed
    latency_seconds: float
    prompt_tokens: int
    completion_tokens: int


@dataclass
class BenchmarkReport:
    """Aggregate benchmark results."""

    results: list[EvalResult] = field(default_factory=list)
    total_questions: int = 0
    correct_count: int = 0
    total_latency: float = 0.0
    total_prompt_tokens: int = 0
    total_completion_tokens: int = 0

    @property
    def accuracy(self) -> float:
        if self.total_questions == 0:
            return 0.0
        return self.correct_count / self.total_questions

    @property
    def mean_latency(self) -> float:
        if self.total_questions == 0:
            return 0.0
        return self.total_latency / self.total_questions

    @property
    def mean_prompt_tokens(self) -> float:
        if self.total_questions == 0:
            return 0.0
        return self.total_prompt_tokens / self.total_questions

    def by_difficulty(self) -> dict[str, dict]:
        """Break down results by difficulty tier."""
        groups: dict[str, list[EvalResult]] = {}
        for r in self.results:
            d = r.question.difficulty
            groups.setdefault(d, []).append(r)

        breakdown = {}
        for diff, results in groups.items():
            correct = sum(1 for r in results if r.correct)
            breakdown[diff] = {
                "count": len(results),
                "correct": correct,
                "accuracy": correct / len(results) if results else 0,
                "mean_latency": sum(r.latency_seconds for r in results) / len(results),
            }
        return breakdown


def load_questions(config: FRTRConfig) -> list[QuestionRecord]:
    """Load all questions from Questions sheets across all workbooks."""
    xlsx_files = sorted(config.data_dir.glob("frtr_*.xlsx"))
    questions = []

    for xlsx_path in xlsx_files:
        wb = load_workbook(str(xlsx_path), data_only=True)
        workbook_name = xlsx_path.stem

        # Find Questions sheet
        q_sheet = None
        for name in wb.sheetnames:
            if name.lower() == "questions":
                q_sheet = wb[name]
                break

        if q_sheet is None:
            wb.close()
            continue

        rows = list(q_sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            wb.close()
            continue

        # Parse header
        headers = [str(h).lower().strip() if h else "" for h in rows[0]]

        for row in rows[1:]:
            if not row or not row[0]:
                continue

            values = list(row) + [None] * (len(headers) - len(row))

            q_text = str(values[0]) if values[0] else ""
            if not q_text.strip():
                continue

            reasoning_type = str(values[1]) if len(values) > 1 and values[1] else ""
            answer = str(values[2]) if len(values) > 2 and values[2] else ""
            provenance = str(values[3]) if len(values) > 3 and values[3] else ""
            difficulty = str(values[4]) if len(values) > 4 and values[4] else "Medium"

            questions.append(
                QuestionRecord(
                    workbook_name=workbook_name,
                    workbook_path=xlsx_path,
                    question=q_text,
                    reasoning_type=reasoning_type,
                    expected_answer=answer,
                    provenance=provenance,
                    difficulty=difficulty,
                )
            )

        wb.close()

    return questions


def _extract_images_from_workbook(workbook_path: Path) -> dict[str, bytes]:
    """Extract embedded images from a workbook for passing to the LLM."""
    images = {}
    wb = load_workbook(str(workbook_path))
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for img_idx, img_obj in enumerate(ws._images):
            try:
                img_data = img_obj._data()
                image_id = f"{sheet_name}_image_{img_idx + 1:03d}"
                images[image_id] = img_data
            except Exception:
                pass
    wb.close()
    return images


def _check_answer_match(predicted: str, expected: str) -> Optional[bool]:
    """Compare predicted answer to ground truth.

    Section 4.2: Answers are considered correct if numerically consistent
    or functionally identical in formula logic.

    Returns True/False for clear matches, None for uncertain cases
    requiring manual review.
    """
    pred = predicted.strip().lower()
    exp = expected.strip().lower()

    # Direct string match
    if pred == exp:
        return True

    # Strip "see " prefix from expected (e.g., "See Summary!B4")
    if exp.startswith("see "):
        exp_clean = exp[4:].strip()
        if pred == exp_clean:
            return True
        # Check if the cell reference appears in the prediction
        if exp_clean in pred:
            return True

    # Check if expected is contained in predicted (cell refs)
    if exp in pred or pred in exp:
        return True

    # Numeric comparison
    try:
        pred_num = float(pred.replace(",", "").replace("$", "").replace("%", ""))
        exp_num = float(exp.replace(",", "").replace("$", "").replace("%", ""))
        if abs(pred_num - exp_num) < 0.01 * max(abs(exp_num), 1):
            return True
    except ValueError:
        pass

    # For trend/visual questions, check keyword overlap
    trend_keywords = ["increasing", "decreasing", "declining", "peaked", "yes", "no"]
    if any(kw in pred for kw in trend_keywords) and any(kw in exp for kw in trend_keywords):
        # Both contain trend language - check agreement
        if ("yes" in pred and "yes" in exp) or ("no" in pred and "no" in exp):
            return True
        if ("increas" in pred and "increas" in exp) or ("declin" in pred and "declin" in exp):
            return True

    # Cannot determine automatically
    return None


class Evaluator:
    """Benchmark evaluator that runs the full FRTR pipeline."""

    def __init__(
        self,
        retriever: HybridRetriever,
        reasoner: LLMReasoner,
        config: FRTRConfig,
    ) -> None:
        self._retriever = retriever
        self._reasoner = reasoner
        self._config = config

    def evaluate_question(self, question: QuestionRecord) -> EvalResult:
        """Run the FRTR pipeline on a single question and evaluate."""
        # Stage 2: Retrieve
        chunks = self._retriever.retrieve(
            query=question.question,
            workbook_filter=question.workbook_name,
        )

        # Load images for this workbook
        image_data = _extract_images_from_workbook(question.workbook_path)

        # Stage 3: Reason
        result = self._reasoner.reason(
            query=question.question,
            chunks=chunks,
            image_data=image_data,
        )

        # Evaluate
        correct = _check_answer_match(result.answer, question.expected_answer)

        return EvalResult(
            question=question,
            predicted_answer=result.answer,
            reasoning=result.reasoning,
            correct=correct,
            latency_seconds=result.latency_seconds,
            prompt_tokens=result.prompt_tokens,
            completion_tokens=result.completion_tokens,
        )

    def run_benchmark(
        self,
        questions: Optional[list[QuestionRecord]] = None,
        max_questions: Optional[int] = None,
    ) -> BenchmarkReport:
        """Run the full benchmark evaluation.

        Args:
            questions:     Questions to evaluate (loads all if None)
            max_questions: Limit number of questions (for testing)
        """
        if questions is None:
            questions = load_questions(self._config)

        if max_questions:
            questions = questions[:max_questions]

        report = BenchmarkReport(total_questions=len(questions))

        for i, q in enumerate(questions):
            console.print(
                f"[bold cyan][{i+1}/{len(questions)}][/bold cyan] "
                f"{q.workbook_name}: {q.question[:80]}..."
            )

            try:
                result = self.evaluate_question(q)
                report.results.append(result)

                if result.correct is True:
                    report.correct_count += 1
                    status = "[green]CORRECT[/green]"
                elif result.correct is False:
                    status = "[red]WRONG[/red]"
                else:
                    status = "[yellow]UNCERTAIN[/yellow]"

                report.total_latency += result.latency_seconds
                report.total_prompt_tokens += result.prompt_tokens
                report.total_completion_tokens += result.completion_tokens

                console.print(
                    f"  {status} | "
                    f"Predicted: {result.predicted_answer[:60]} | "
                    f"Expected: {q.expected_answer[:60]} | "
                    f"Latency: {result.latency_seconds:.2f}s"
                )
            except Exception as e:
                console.print(f"  [red]ERROR: {e}[/red]")

        return report


def print_report(report: BenchmarkReport) -> None:
    """Print a formatted benchmark report matching the paper's metrics."""
    console.print("\n" + "=" * 70)
    console.print("[bold]FRTR-Bench Evaluation Report[/bold]")
    console.print("=" * 70)

    # Overall metrics (Section 4.2)
    table = Table(title="Overall Metrics")
    table.add_column("Metric", style="bold")
    table.add_column("Value", justify="right")
    table.add_row("Total Questions", str(report.total_questions))
    table.add_row("Correct", str(report.correct_count))
    table.add_row("Answer Accuracy", f"{report.accuracy:.2%}")
    table.add_row("Mean Latency (s)", f"{report.mean_latency:.2f}")
    table.add_row("Mean Prompt Tokens", f"{report.mean_prompt_tokens:.0f}")
    console.print(table)

    # By difficulty (Table 2 / Figure 2)
    breakdown = report.by_difficulty()
    if breakdown:
        diff_table = Table(title="Results by Difficulty")
        diff_table.add_column("Difficulty", style="bold")
        diff_table.add_column("Count", justify="right")
        diff_table.add_column("Correct", justify="right")
        diff_table.add_column("Accuracy", justify="right")
        diff_table.add_column("Mean Latency", justify="right")

        for diff in ["Easy", "Medium", "Hard"]:
            if diff in breakdown:
                d = breakdown[diff]
                diff_table.add_row(
                    diff,
                    str(d["count"]),
                    str(d["correct"]),
                    f"{d['accuracy']:.2%}",
                    f"{d['mean_latency']:.2f}s",
                )
        console.print(diff_table)

    console.print("=" * 70)


def save_report(report: BenchmarkReport, output_path: Path) -> None:
    """Save benchmark results to a JSON file."""
    data = {
        "summary": {
            "total_questions": report.total_questions,
            "correct": report.correct_count,
            "accuracy": report.accuracy,
            "mean_latency_s": report.mean_latency,
            "mean_prompt_tokens": report.mean_prompt_tokens,
            "total_prompt_tokens": report.total_prompt_tokens,
            "total_completion_tokens": report.total_completion_tokens,
        },
        "by_difficulty": report.by_difficulty(),
        "results": [
            {
                "workbook": r.question.workbook_name,
                "question": r.question.question,
                "reasoning_type": r.question.reasoning_type,
                "difficulty": r.question.difficulty,
                "expected_answer": r.question.expected_answer,
                "predicted_answer": r.predicted_answer,
                "reasoning": r.reasoning,
                "correct": r.correct,
                "latency_s": r.latency_seconds,
                "prompt_tokens": r.prompt_tokens,
            }
            for r in report.results
        ],
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    console.print(f"\nResults saved to {output_path}")

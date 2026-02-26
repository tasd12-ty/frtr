"""Stage 1: Multi-granular workbook decomposition and embedding.

Implements the offline indexing phase of Algorithm 1 in the FRTR paper.
Decomposes Excel workbooks into four unit types:
  - Rows:    each row serialized with column headers
  - Columns: each column serialized with row indices
  - Windows: s×s sliding context windows preserving spatial layout
  - Images:  embedded PNG images extracted from sheets
"""

from __future__ import annotations

import io
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image
from tqdm import tqdm

from .config import FRTRConfig
from .embeddings.base import EmbeddingBackend


@dataclass
class ChunkUnit:
    """A single indexed unit from a workbook."""

    text: str
    unit_type: str  # "row", "column", "window", "image"
    sheet_name: str
    workbook_name: str
    metadata: dict = field(default_factory=dict)
    embedding: Optional[np.ndarray] = None
    image: Optional[Image.Image] = None  # Only for image units


def _cell_value_str(val) -> str:
    """Convert a cell value to a clean string."""
    if val is None:
        return ""
    return str(val)


def _serialize_row(headers: list[str], row_values: list, row_idx: int) -> str:
    """Serialize a row with column headers for context.

    Paper Section 3.4: 'Rows: include column headers'
    Format: ROW_{idx}: Header1=Value1 | Header2=Value2 | ...
    """
    parts = []
    for h, v in zip(headers, row_values):
        v_str = _cell_value_str(v)
        if v_str:
            parts.append(f"{h}={v_str}")
    return f"ROW_{row_idx}: " + " | ".join(parts)


def _serialize_column(col_header: str, col_values: list, row_indices: list[int]) -> str:
    """Serialize a column with row indices for context.

    Paper Section 3.4: 'Columns: include row indices'
    Format: COL_{header}: [row_1]=val1, [row_2]=val2, ...
    """
    parts = []
    for idx, v in zip(row_indices, col_values):
        v_str = _cell_value_str(v)
        if v_str:
            parts.append(f"[{idx}]={v_str}")
    return f"COL_{col_header}: " + ", ".join(parts)


def _serialize_window(
    headers: list[str],
    data: list[list],
    start_row: int,
    start_col: int,
    window_size: int,
) -> str:
    """Serialize an s×s sliding window preserving spatial layout.

    Paper Section 3.4: 'Windows: preserve s×s spatial layout'
    """
    lines = []
    col_headers = headers[start_col : start_col + window_size]
    lines.append(f"WINDOW[{start_row}:{start_row+window_size}, {start_col}:{start_col+window_size}]")
    lines.append("  " + " | ".join(col_headers))
    for r_offset, row in enumerate(data):
        row_slice = row[start_col : start_col + window_size]
        vals = [_cell_value_str(v) for v in row_slice]
        lines.append(f"  R{start_row + r_offset}: " + " | ".join(vals))
    return "\n".join(lines)


def decompose_sheet(
    ws: Worksheet,
    sheet_name: str,
    workbook_name: str,
    k_target: int,
) -> list[ChunkUnit]:
    """Decompose a single worksheet into row, column, and window chunks.

    Algorithm 1, lines 3-9:
      For each sheet S in W:
        s ← ceil(sqrt(N / K_target))
        For each unit u in {rows, columns, s×s windows, images}: ...
    """
    chunks: list[ChunkUnit] = []

    # Read all data
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows or len(all_rows) < 2:
        return chunks

    # Extract headers from first row
    headers = [_cell_value_str(h) or f"Col{i}" for i, h in enumerate(all_rows[0])]
    data_rows = all_rows[1:]  # Skip header row
    n_rows = len(data_rows)
    n_cols = len(headers)

    if n_rows == 0:
        return chunks

    # --- Row chunks ---
    for i, row in enumerate(data_rows):
        text = _serialize_row(headers, list(row), i + 2)  # 1-indexed, skip header
        if text.strip() and text != f"ROW_{i+2}: ":
            chunks.append(
                ChunkUnit(
                    text=text,
                    unit_type="row",
                    sheet_name=sheet_name,
                    workbook_name=workbook_name,
                    metadata={"row_index": i + 2, "headers": headers},
                )
            )

    # --- Column chunks ---
    for col_idx in range(n_cols):
        col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
        non_empty = [v for v in col_values if v is not None and str(v).strip()]
        if non_empty:
            row_indices = list(range(2, n_rows + 2))
            text = _serialize_column(headers[col_idx], col_values, row_indices)
            chunks.append(
                ChunkUnit(
                    text=text,
                    unit_type="column",
                    sheet_name=sheet_name,
                    workbook_name=workbook_name,
                    metadata={"col_index": col_idx, "col_header": headers[col_idx]},
                )
            )

    # --- Sliding window chunks ---
    # s = ceil(sqrt(N / K_target)) per Algorithm 1 line 4
    total_cells = n_rows * n_cols
    s = max(2, math.ceil(math.sqrt(total_cells / max(k_target, 1))))

    for r_start in range(0, n_rows, s):
        for c_start in range(0, n_cols, s):
            window_rows = data_rows[r_start : r_start + s]
            if not window_rows:
                continue
            text = _serialize_window(
                headers, window_rows, r_start + 2, c_start, s
            )
            # Only add non-trivial windows
            if any(
                _cell_value_str(v).strip()
                for row in window_rows
                for v in row[c_start : c_start + s]
                if v is not None
            ):
                chunks.append(
                    ChunkUnit(
                        text=text,
                        unit_type="window",
                        sheet_name=sheet_name,
                        workbook_name=workbook_name,
                        metadata={
                            "row_start": r_start + 2,
                            "col_start": c_start,
                            "window_size": s,
                        },
                    )
                )

    # --- Image chunks ---
    for img_idx, img_obj in enumerate(ws._images):
        try:
            img_data = img_obj._data()
            pil_image = Image.open(io.BytesIO(img_data)).convert("RGB")
            image_id = f"{sheet_name}_image_{img_idx + 1:03d}"
            chunks.append(
                ChunkUnit(
                    text=f"[Image: {image_id}] Embedded image from sheet '{sheet_name}'",
                    unit_type="image",
                    sheet_name=sheet_name,
                    workbook_name=workbook_name,
                    metadata={"image_id": image_id},
                    image=pil_image,
                )
            )
        except Exception as e:
            print(f"Warning: Could not extract image {img_idx} from {sheet_name}: {e}")

    return chunks


def index_workbook(
    workbook_path: Path,
    embedder: EmbeddingBackend,
    config: FRTRConfig,
) -> list[ChunkUnit]:
    """Index a single Excel workbook.

    Decomposes all non-metadata/non-question sheets into chunks,
    embeds each chunk, and returns the list of ChunkUnits.

    Algorithm 1, lines 1-10.
    """
    wb = load_workbook(str(workbook_path), data_only=True)
    workbook_name = workbook_path.stem
    all_chunks: list[ChunkUnit] = []

    skip_sheets = {"questions", "readme"}

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in skip_sheets:
            continue

        ws = wb[sheet_name]
        chunks = decompose_sheet(ws, sheet_name, workbook_name, config.k_target)
        all_chunks.extend(chunks)

    wb.close()

    # Embed all chunks
    if all_chunks:
        # Separate text and image chunks
        text_chunks = [c for c in all_chunks if c.unit_type != "image"]
        image_chunks = [c for c in all_chunks if c.unit_type == "image"]

        # Batch embed text chunks
        if text_chunks:
            texts = [c.text for c in text_chunks]
            batch_size = 256
            embeddings_list = []
            for i in range(0, len(texts), batch_size):
                batch = texts[i : i + batch_size]
                batch_emb = embedder.embed_texts(batch)
                embeddings_list.append(batch_emb)
            all_embeddings = np.concatenate(embeddings_list, axis=0)
            for chunk, emb in zip(text_chunks, all_embeddings):
                chunk.embedding = emb

        # Embed images individually
        for chunk in image_chunks:
            if chunk.image is not None:
                chunk.embedding = embedder.embed_image(chunk.image)

    return all_chunks


def index_all_workbooks(
    config: FRTRConfig,
    embedder: EmbeddingBackend,
) -> list[ChunkUnit]:
    """Index all .xlsx workbooks in the data directory.

    Returns a flat list of all ChunkUnits across all workbooks.
    """
    xlsx_files = sorted(config.data_dir.glob("frtr_*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No frtr_*.xlsx files found in {config.data_dir}")

    all_chunks: list[ChunkUnit] = []
    for xlsx_path in tqdm(xlsx_files, desc="Indexing workbooks"):
        chunks = index_workbook(xlsx_path, embedder, config)
        all_chunks.extend(chunks)
        tqdm.write(f"  {xlsx_path.name}: {len(chunks)} chunks")

    return all_chunks
